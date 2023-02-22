# Модуль, который сканирует бланки в папке бланков и создает из них процедуры

import os, shutil, time, datetime, copy
from openpyxl import load_workbook, Workbook
from specparser import SpecParser
from aux_code import load_first_sheet


class BlanksParser:
    def __init__(self, xl_path, params: dict):
        self.path = xl_path
        self.sheet = self._load_first_sheet()
        self.pp_number = params.get('PP_NUMBER')
        self.pp_header = params.get('PP_HEADER')
        if not self.is_pp():
            raise FileNotFoundError('Данный файл не является файлом ПП-бланка:', self.path)

    def _load_first_sheet(self):
        return load_first_sheet(self.path)

    # Метод, который анализирует xlsx-файл по данному пути и возвращает dict c парами заголовок-фильтр
    def get_filters(self):
        forming_dict = dict()
        for j in range(1, self.sheet.max_column):
            header = self.sheet.cell(row=3, column=j).value
            subfilter = self.sheet.cell(row=4, column=j).value
            if header == self.pp_number:
                continue
            elif header:                              # Чтобы None не превращалось в 'none'
                header = str(header)
            else:
                header = ''
            if subfilter:                           # Чтобы None не превращалось в 'none'
                subfilter = str(subfilter).lower()
            else:
                subfilter = ''
            forming_dict.update({header: subfilter})
        if not forming_dict:
            raise FileNotFoundError('Ошибка формирования словаря фильтров: ' + str(self.path))
        return forming_dict

    # Проверяет xlsx-файлы на соответствие критериям
    # Сейчас это: первая ячейка первого листа должна содержать XL_HEADER
    def is_pp(self):
        if self.sheet.cell(row=1, column=1).value == self.pp_header:
            return True
        return None

    def get_copy(self):
        new_wb = Workbook()
        new_ws = new_wb.active
        for row in self.sheet:
            line = list()
            for cell in row:
                line.append(cell.value)
            new_ws.append(line)
        return new_ws


# Класс BlankHandler, занимающийся сбором и анализом xlsx-бланков заказов
class BlanksHandler:
    # При инициализации объекта класса, нужно скормить ему объект класса ConfigParser модуля сonfigparser,
    # инициализированный через ConfigParser.read('path')
    def __init__(self, params: dict):
        self.search_range = params.get("blanks_path")
        self.PP_HEADER = params.get("PP_HEADER")
        self.params = params
        self.found_blanks = self.get_blanks_paths()

    # Ищет файлы бланков в директории, прописанной в конфиге
    def get_blanks_paths(self):
        blanks_pathlist = []
        with os.scandir(path=self.search_range) as directory:
            for entry in directory:
                if ".xlsx" in entry.name and not '~$' in entry.name:
                    joint_path = '/'.join((self.search_range, entry.name))
                    parser = BlanksParser(joint_path, self.params)
                    if parser.is_pp():
                        blanks_pathlist.append(joint_path)
        if len(blanks_pathlist) == 0:
            raise FileNotFoundError("Файлов бланков не найдено!")
        return blanks_pathlist

    def save_formatted_copy(self, index: int, path_destination: str):
        bparser = BlanksParser(self.found_blanks[index], self.params)
        new_wb = bparser.get_copy().parent
        new_ws = self.format_sheet(new_wb.active)
        new_wb.save(path_destination)
        bparser = BlanksParser(path_destination, self.params)
        print(bparser.get_filters())

    def format_sheet(self, sheet):
        sheet.delete_rows(1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
        for row in sheet:
            for cell in row:
                if cell.value:
                    for key, value in self.params.items():
                        cell.value = str(cell.value).replace(key, value)
        return sheet



    def fill_blank(self, path):
        # parser =
        pass

    def form_blank(self, path_source, path_destination):
        pass    # копировать шаблон
        pass    # наполнить его
        pass    # save it


class BlankFiller:
    def __init__(self, filters: dict, partslist: list[dict], params: dict, xl_path='', sheet=None):
        if xl_path != '':
            self.blank = load_first_sheet(xl_path)
        elif not sheet:
            self.blank = sheet
        else:
            raise RuntimeError('Ошибка заполнения бланка')
        self.partslist = partslist
        self.filters = filters
        self.sheet = sheet

    def get_matches(self):
        ls_out = list()
        for entry in self.partslist:
            flag1 = True
            entry_out = dict()
            for key, value in self.filters.items():
                if key in entry.keys():
                    if value in entry[key]:
                        print(key, entry[key])
                        entry_out.update({key: entry[key]})
                    else:
                        continue
                else:
                    continue
            # if entry_out != {}:
            ls_out.append(entry_out)
        return ls_out


# Для тестов
if __name__ == '__main__':
    from configparser import ConfigParser
    configur = ConfigParser()
    configur.read('config.ini')
    configur_now = {'blanks_path': configur.get("paths", "blanks_path"),
                    'PP_HEADER': configur.get("variables", "PP_HEADER"),
                    '<PRODUCT>': 'Стеллаж',
                    '<AUTHOR>': 'Пар-оол Ч.А.',
                    '<DOCS_FOLDER>': configur.get("common folders", "folder_2"),
                    '<DATE_OF_ISSUE>': str(datetime.date.today()),
                    '<OTHER>': '',
                    'PP_NUMBER': 'Номер',
                    'POSITION_MARKER': 'Поз.',
                    'QUANTITY_MARKER': 'Кол-во'
                    }
    bhandler = BlanksHandler(configur_now)
    # creating new filename
    old_path = bhandler.get_blanks_paths()[0]
    # new_ls = old_path.split('.')
    # new_ls[1] = new_ls[1] + ' ' + '_'.join((str(time.gmtime().tm_hour), str(time.gmtime().tm_min), str(time.gmtime().tm_sec)))
    # new_path = '.'.join(new_ls)
    # copying
    # bhandler.save_formatted_copy(0, new_path)
    bparser = BlanksParser(old_path, configur_now)
    sparser = SpecParser('./Пример спеки/Shifted spec.xlsx', configur_now)
    bfiller = BlankFiller(bparser.get_filters(), sparser.get_flat_unformatted(), configur_now)
    for enrty in bfiller.get_matches():
        print( )

    # print(bparser.get_filters())
    # bfiller = BlankFiller(new_path)
