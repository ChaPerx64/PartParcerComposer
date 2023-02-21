from openpyxl import load_workbook

# SPECS = 'specs'

def p(what='foo'):
    print(what)


class SpecParser:
    def __init__(self, xl_path, key_pos, key_qty):
        self.path = xl_path
        self.sheet = self._load_first_sheet()
        self.pos_key = key_pos.lower()
        self.qty_key = key_qty.lower()
        self.specs_key = 'specs'

    def _load_first_sheet(self, remove_enters=True):
        try:
            xl_workbook = load_workbook(self.path, read_only=False)
        except:
            raise RuntimeError('Ошибка чтения xlsx-файла')
        active_sheet = xl_workbook[xl_workbook.sheetnames[0]]
        for row in active_sheet.rows:
            for cell in row:
                if remove_enters:
                    if cell.value:
                        new_value = str(cell.value).replace('\n', '')
                        cell.value = new_value
        return active_sheet

    # Метод, получающий заголовки из листа
    def _get_spec_headers(self):
        headers = []
        for row in self.sheet.rows:
            for cell in row:
                if cell.value:
                    headers.append(str(cell.value).lower())
            break
        return headers

    def _get_spec_rows(self):
        rows = []
        count = 1
        for row in self.sheet.rows:
            line = []
            if count > 1:
                for cell in row:
                    if cell.value:
                        line.append(str(cell.value))
                    else:
                        line.append(None)
                rows.append(line)
            count += 1
        return rows

    def get_entries_raw(self):
        entries = list()
        headers = self._get_spec_headers()
        rows = self._get_spec_rows()
        for row in rows:
            entry = dict()
            for header, value in zip(headers, row):
                entry.update({header: value})
            entries.append(entry)
        return entries

    def entry_to_dict(self, entry: dict):
        out_dict = dict()
        specs_dict = dict()
        for key, value in entry.items():
            if key == self.pos_key:
                out_dict.update({key: value})
            elif key == self.qty_key:
                out_dict.update({key: value})
            else:
                specs_dict.update({key: value})
        out_dict.update({self.specs_key: specs_dict})
        return out_dict

    def get_entries(self):
        entries = list()
        for entry in self.get_entries_raw():
            entries.append(self.entry_to_dict(entry))
        return entries

    def get_counted(self):
        prev_depth = 0
        prev_qty = 0
        mplicators_ls = [1]
        counted_list = list()
        for entry in self.get_entries():
            qty = int(entry[self.qty_key])
            depth = str(entry[self.pos_key]).count('.')
            if prev_depth < depth:
                mplicators_ls.append(prev_qty)
            if prev_depth > depth:
                mplicators_ls.pop(len(mplicators_ls)-1)
            mplicator = 1
            for a in mplicators_ls:
                mplicator = mplicator * a
            entry[self.qty_key] = str(qty * mplicator)
            counted_list.append(entry)
            prev_qty = qty
            prev_depth = depth
        return counted_list

    def get_flatlist(self):
        flat_list = list()
        counted = False
        for entry in self.get_counted():
            for fl_entry in flat_list:
                if fl_entry[self.specs_key] == entry[self.specs_key]:
                    fl_entry[self.qty_key] = str(int(entry[self.qty_key]) + int(fl_entry[self.qty_key]))
                    counted = True
                    break
            if not counted:
                flat_list.append(entry)
        return flat_list

    def dict_to_entry(self, dict_entry:dict):
        for key, value in dict_entry[self.specs_key].items():
            dict_entry.update({key: value})
        dict_entry.pop(self.specs_key)
        return dict_entry

    def get_flat_unformatted(self):
        out_list = list()
        for entry in self.get_flatlist():
            out_list.append(self.dict_to_entry(entry))
        return out_list



# for tests
XL_PATH = './Пример спеки/Shifted spec.xlsx'
parser = SpecParser(XL_PATH, 'Поз.', 'Кол-во')
count = 1
for item in parser.get_flat_unformatted():
    print('\nEntry #' + str(count))
    print(item)
    # for key, value in item.items():
    #     print('     ', key, ': ', value, sep='')
    count += 1

# Функция, вытаскивающая детали из листа эксель, и распределяющая их по спискам
# def parts_list_exctractor(worksheet, laser_list: list, tokar_list: list):
#     coord = file_checker(worksheet)
#     if coord is None:
#         return laser_list, tokar_list
#     i = 0
#     parts_list = list()
#     for row in worksheet.rows:
#         if i < coord['lstart']:
#             i += 1
#             continue
#         part_prop = dict()
#         for item in COL_NAMES:
#             part_prop.update({item: row[coord[item]].value})
#         parts_list.append(part_prop)
#     for part in parts_list:
#         if not part['marsh'] is None:
#             if 'Лазер' in str(part['marsh']):
#                 laser_list = add_part_to_list(laser_list, part)
#             if 'Токар' in str(part['marsh']):
#                 tokar_list = add_part_to_list(tokar_list, part)
#     return laser_list, tokar_list
