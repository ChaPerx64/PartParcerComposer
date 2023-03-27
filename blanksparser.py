from aux_code import load_first_sheet
from openpyxl import Workbook

class BlanksParser:
    """Constructs a new instance of BlanksParser taking a path to the PP-blank xlsx-file"""
    def __init__(self, xl_path, params: dict):
        self.path = xl_path
        self.sheet = self._load_first_sheet()
        self.pp_number = params.get('NUMBER_MARKER')
        self.pp_header = params.get('PP_HEADER')

    def _load_first_sheet(self):
        return load_first_sheet(self.path)

    """Метод, который анализирует xlsx-файл по данному пути и возвращает dict c парами заголовок-фильтр"""
    def get_filters(self):
        if not self.is_pp():
            raise FileNotFoundError('Данный файл не является файлом ПП-бланка:', self.path)
        forming_dict = dict()
        for j in range(1, self.sheet.max_column+1):
            header = self.sheet.cell(row=3, column=j).value
            filter_str = self.sheet.cell(row=4, column=j).value
            if header == self.pp_number:
                continue
            elif header:
                header = str(header)
            else:
                continue
            filter_ls = list()
            if filter_str:
                for subfilter in str(filter_str).split(','):
                    if subfilter == '*':
                        filter_ls.append('')
                    else:
                        filter_ls.append(str(filter_str).strip())
            else:
                filter_ls = None
            forming_dict.update({header: filter_ls})
        if not forming_dict:
            raise FileNotFoundError('Ошибка формирования словаря фильтров: ' + str(self.path))
        return forming_dict

    """Проверяет xlsx-файлы на соответствие критериям
    Сейчас это: первая ячейка первого листа должна содержать XL_HEADER"""
    def is_pp(self):
        if self.sheet.cell(row=1, column=1).value == self.pp_header:
            return True
        return None

    """Returns a copy instance of parsed Worksheet (in a new Workbook instance).
    No structure (Tables, meerged cells etc.) is retained. """
    def get_copy(self):
        if not self.is_pp():
            raise FileNotFoundError('Данный файл не является файлом ПП-бланка:', self.path)
        new_wb = Workbook()
        new_ws = new_wb.active
        for row in self.sheet:
            line = list()
            for cell in row:
                line.append(cell.value)
            new_ws.append(line)
        return new_ws