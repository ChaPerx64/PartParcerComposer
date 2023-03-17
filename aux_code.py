# Модуль с допфункциями, помогающими работать с эксель
from openpyxl import load_workbook
import datetime


def load_first_sheet(xl_path, remove_enters=True):
    try:
        xl_workbook = load_workbook(xl_path, read_only=False, data_only=True)
    except:
        raise RuntimeError('Ошибка чтения файла:', xl_path)
    active_sheet = xl_workbook[xl_workbook.sheetnames[0]]
    for row in active_sheet.rows:
        for cell in row:
            if remove_enters:
                if cell.value:
                    new_value = str(cell.value).replace('\n', '')
                    cell.value = new_value
    return active_sheet


def filename_from_path(path: str):
    try:
        slices = path.split('.')
        slices = str(slices[-2]).split('/')
        return slices[-1]
    except Exception:
        raise FileNotFoundError


def new_filepath(project_name: str, save_folder: str, blankname: str):
    out_filepath = save_folder + project_name + '   ' + blankname + '.xlsx'
    return out_filepath


def curr_dt_subfolder(orig_path: str):
    orig_slices = orig_path.split('/')
    orig_slices.pop(len(orig_slices)-1)
    new_path = str(datetime.datetime.now()).rsplit('.', 1)[0]
    new_path = new_path.replace('-', '_')
    new_path = new_path.replace(':', '_')
    new_path = new_path.replace(' ', '__')
    new_path = 'Заказы ' + new_path
    new_path = '/'.join(orig_slices + [new_path]) + '/'
    return new_path


def key_from_listitem():
    pass