# This is a sample Python script.
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import datetime
import os
import shutil

from openpyxl import load_workbook

COL_NAMES = ('obozn', 'naimen', 'kolvo', 'razd', 'marsh')
COL_NAMES_SHEET = ('обозначение', "наименование", "кол-во", "раздел", "маршрут изготовления")
LOG_FILENAME = 'log.txt'


class BColors:
    WARNING = '\u001b[31m'
    GREEN = '\u001b[32m'
    YELLOW = '\u001b[33m'
    ENDC = '\033[0m'


# Функция, которая ищет один единственный .xlsx-файл в корневом каталоге
def spec_finder():
    xlsx_filename = ''
    with os.scandir() as center:
        for entry in center:
            if ".xlsx" in entry.name:
                xlsx_filename = entry.name
                break
    if not xlsx_filename:
        screen_holder(BColors.WARNING + 'ОШИБКА: Файл xlsx не найден!' + BColors.ENDC)
        raise FileNotFoundError("Файл xlsx не найден!")
    return xlsx_filename


# Функция, вытаскивающая детали из листа эксель, и распределяющая их по спискам
def parts_list_exctractor(worksheet, laser_list: list, tokar_list: list):
    coord = file_checker(worksheet)
    if coord is None:
        return laser_list, tokar_list
    i = 0
    parts_list = list()
    for row in worksheet.rows:
        if i < coord['lstart']:
            i += 1
            continue
        part_prop = dict()
        for item in COL_NAMES:
            part_prop.update({item: row[coord[item]].value})
        parts_list.append(part_prop)
    for part in parts_list:
        if not part['marsh'] is None:
            if 'Лазер' in str(part['marsh']):
                laser_list = add_part_to_list(laser_list, part)
            if 'Токар' in str(part['marsh']):
                tokar_list = add_part_to_list(tokar_list, part)
    return laser_list, tokar_list


def add_part_to_list(partlist: list, new_part: dict):
    for part in partlist:
        if part[COL_NAMES[0]] == new_part:
            part['kolvo'] = int(part['kolvo']) + int(new_part['kolvo'])
            return partlist
    partlist.append(new_part)
    return partlist


# Функция, котороая ищет строку заголовков с обязательными заголовками
# В случае неудачи выдает ошибку
def file_checker(worksheet):
    res_dict = dict()
    i = 0
    for row in worksheet.rows:
        fl_detected = True
        j = 0
        for cell in row:
            k = 0
            for col in COL_NAMES_SHEET:
                if str(cell.value).lower() == col:
                    res_dict.update({COL_NAMES[k]: j})
                k += 1
            j += 1
        for item in COL_NAMES:
            if res_dict.get(item) is None:
                fl_detected = False
                break
        i += 1
        if fl_detected:
            res_dict.update({'lstart': i})
            return res_dict
    return None


# Функция, которая создает список файлов с заданным расширением из лежащих в корневой папке и подпапках
def file_scanner(extention: str):
    ls_file_ext = list()
    for root, dirs, files in os.walk('.', topdown=True):
        for filename in files:
            try:
                file_title, file_ext = filename.rsplit('.', 1)
                if file_ext.lower() == extention.lower():
                    ls_file_ext.append((file_title, os.path.join(root, filename)))
            except:
                pass
    print('В данном каталоге найдено ' + str(len(ls_file_ext)) + ' ' + extention + '-файлов.')
    return ls_file_ext


# Функция, которая ищет доки в листе доков для соответствующих деталей из списка деталей
# и возвращает список путей к этим докам, прицепив поле с заданным ключом в исходном словаре
def matchmaker(partslist: list, doclist: list, ext: str, lookformain=False):
    ext = ext.lower()
    endline = ' ' + ext + '.'
    for part in partslist:
        pathlist = list()
        for doc in doclist:
            if doc[0] == part[COL_NAMES[0]]:
                pathlist.append(doc[1])

        # ПЕРВАЯ ОЦЕНКА КОЛИЧЕСТВА НАЙДЕННЫХ ФАЙЛОВ ДЛЯ ПОИСКА ГРУППОВЫХ ЧЕРТЕЖЕЙ
        ls_length = len(pathlist)
        if ls_length == 0 and '-' in part[COL_NAMES[0]]:
            print(BColors.YELLOW + 'Для детали ' + part[COL_NAMES[0]] + ' не найдено' + endline + BColors.ENDC)
            if lookformain:
                print('Пробую найти групповой чертёж...')
                groupdocname = str(part[COL_NAMES[0]]).rsplit('-', 1)[0]
                log_out_txt(
                    'Для детали ' + part[COL_NAMES[0]] + ' выполнялся поиск группового чертежа по обозначению: ' +
                    groupdocname + '\n'
                )
                for doc in doclist:
                    if doc[0] == groupdocname:
                        pathlist.append(doc[1])
            else:
                print('Поиск группового чертежа не выполнялся')
                log_out_txt(
                    'Для детали ' + part[COL_NAMES[0]] + ' поиск группового чертежа не выполнялся\n'
                )


        # ОКОНЧАТЕЛЬНАЯ ОЦЕНКА КОЛИЧЕСТВА НАЙДЕННЫХ ФАЙЛОВ
        ls_length = len(pathlist)
        if ls_length == 0:
            print(BColors.WARNING + 'Для детали ' + part[COL_NAMES[0]] + ' не найдено' + endline + BColors.ENDC)
            part.update({ext: None})
        elif ls_length == 1:
            print(
                BColors.GREEN + 'Для детали ' + part[
                    COL_NAMES[0]] + ' обнаружен единственный' + endline + BColors.ENDC +
                '\nВыбран файл из директории: ' + pathlist[0]
            )
            part.update({ext: pathlist[0]})
        else:
            youngest_file_datetime = 0
            for path in pathlist:
                if os.path.getmtime(path) > youngest_file_datetime:
                    youngest_file_path = path
                    youngest_file_datetime = os.path.getmtime(path)
            part.update({ext: youngest_file_path})
            print(
                BColors.YELLOW + part[COL_NAMES[0]] + '  -----  Обнаружено ' + str(ls_length) + endline + BColors.ENDC +
                '\nВыбран файл из директории: ' + youngest_file_path +
                '\nИзменён: ' + str(datetime.datetime.fromtimestamp(youngest_file_datetime))
            )
    return partslist


def curr_dt_folder():
    new_path = str(datetime.datetime.now()).rsplit('.', 1)[0]
    new_path = new_path.replace('-', '_')
    new_path = new_path.replace(':', '_')
    new_path = new_path.replace(' ', '__')
    new_path = './' + new_path + '/'
    return new_path


def files_copier(partslist: list, ext: str, subpath: os.PathLike):
    if len(partslist) == 0:
        return
    for part in partslist:
        s_path = part[ext]
        if s_path is not None:
            os.makedirs(subpath, exist_ok=True)
            shutil.copy2(s_path, subpath)
        else:
            log_out_txt(
                part[COL_NAMES[0]] + '  -----  ' +
                ext.upper() + '-ФАЙЛ НЕ НАЙДЕН\n'
            )
            print(
                BColors.WARNING + part[COL_NAMES[0]] + '  -----  ' +
                ext.upper() + '-ФАЙЛ НЕ НАЙДЕН' + BColors.ENDC
            )


def log_out_txt(text: str):
    with open(LOG_FILENAME, 'a') as f:
        f.write(text)


def screen_holder(text: str):
    if text:
        print(text)
    print('Нажми "Enter", чтобы закрыть окно')
    input()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print('KD_Crab 0.1 welcomes you!')
    print('\nИщу спецификацию...')
    speca = load_workbook(spec_finder(), read_only=True)
    print('Обнаружен файл ' + spec_finder() + '.\n')

    # СОЗДАНИЕ СПИСКОВ ДЕТАЛЕЙ НА ТОКАРКУ И ЛАЗЕРКУ ИЗ РАБОЧИХ ЛИСТОВ
    ls_laser = []
    ls_tokar = []
    for sheetname in speca.sheetnames:
        active_ws = speca[sheetname]
        ls_laser, ls_tokar = parts_list_exctractor(active_ws, ls_laser, ls_tokar)
        # if ls_laser is None or ls_tokar is None:
        #     continue
        log_out_txt(
            '\nЛист "' + sheetname + '"\n'
        )
        print('Читаю лист ' + sheetname)

    if len(ls_laser) != 0:
        print("\nСписок деталей, отмеченных для лазера:")
        for item in ls_laser:
            for value in item.values():
                print(value, end='              ')
            print()
    if len(ls_tokar) != 0:
        print("\nСписок деталей, отмеченных для токарки:")
        for item in ls_tokar:
            for value in item.values():
                print(value, end='              ')
            print()

    # ОБРАБОТКА ОШИБКИ ПУСТОГО ИЛИ НЕКОРРЕКТНОГО XLSX ФАЙЛА
    if len(ls_laser) == 0 and len(ls_tokar) == 0:
        screen_holder('Найденный файл не является файлом спецификации или неверно отформатирован')
        raise FileNotFoundError('Найденный файл не является файлом спецификации или неверно отформатирован')

    # СКАНИРОВАНИЕ ДИРЕКТОРИИ НА PDF И DXF ФАЙЛЫ
    print('\nИщу документы, соответствующие обозначению...')
    pdf_files = file_scanner('pdf')
    dxf_files = file_scanner('dxf')
    if len(pdf_files) == 0 or (len(ls_laser) != 0 and len(dxf_files) == 0):
        screen_holder(
            BColors.WARNING + 'ОШИБКА: Данная директория не содержит файлы pdf или dxf' + BColors.ENDC
        )
        raise FileNotFoundError('Данная директория не содержит файлы pdf или dxf')

    # СОЗДАНИЕ ИМЕНИ НОВОЙ ДИРЕКТОРИИ ДЛЯ НАЙДЕННЫХ ФАЙЛОВ
    new_folder_path = curr_dt_folder()
    # logfile_path = new_folder_path + '.txt'
    if os.path.exists("log.txt"):
        os.remove("log.txt")

    # ПОИСК ДОКОВ ДЛЯ ТОКАРКИ
    log_out_txt("\nТОКАРКА, PDF\n")
    print('\nИщу КД на детали, помеченные для токарки...')
    ls_tokar = matchmaker(ls_tokar, pdf_files, 'pdf', lookformain=True)
    print('\nКопирую обнаруженные файлы токарки в созданную директорию...')
    files_copier(ls_tokar, 'pdf', os.path.join(new_folder_path, 'Токарка/'))
    print('Копирование завершено')

    # ПОИСК ДОКОВ ДЛЯ ЛАЗЕРА
    log_out_txt("\nЛАЗЕРКА, PDF\n")
    print('\nИщу КД на детали, помеченные для лазера...')
    ls_laser = matchmaker(ls_laser, pdf_files, 'pdf', lookformain=True)
    files_copier(ls_laser, 'pdf', os.path.join(new_folder_path, 'Лазер/PDF/'))
    log_out_txt("\nЛАЗЕРКА, DXF\n")
    ls_laser = matchmaker(ls_laser, dxf_files, 'dxf')
    files_copier(ls_laser, 'dxf', os.path.join(new_folder_path, 'Лазер/DXF/'))



    print('Копирование завершено')

    # ПАУЗА ДЛЯ ПРОЧТЕНИЯ ВСЕГО, ЧТО НАРОДИЛ СКРИПТ :)
    screen_holder(
        BColors.GREEN + '\nВЫПОЛНЕНИЕ СКРИПТА ЗАВЕРШЕНО.' + BColors.ENDC +
        '\nВажные примечания записаны в файл log.txt'
    )

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

